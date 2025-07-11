import time
from openpyxl import load_workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from fuzzywuzzy import process

# === Load Excel with formatting preserved ===
wb = load_workbook("Hhh.xlsx", data_only=False)
ws = wb.active

# Get headers (first row)
headers = [cell.value for cell in ws[1]]

# Extract formatted/display values for each row
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_data = {}
    for header, cell in zip(headers, row):
        if cell.value is None:
            display_value = ""
        elif cell.number_format in ["0%", "0.00%", "##0.0%", "##0%", "0.##%"]:
            try:
                percentage = float(cell.value) * 100
                if percentage.is_integer():
                    display_value = f"{int(percentage)}%"
                else:
                    display_value = f"{percentage:.2f}".rstrip('0').rstrip('.') + '%'
            except:
                display_value = str(cell.value)
        elif cell.is_date:
            display_value = cell.value.strftime('%Y-%m-%d')
        else:
            display_value = str(cell.value)
        row_data[header] = display_value.strip()
    data.append(row_data)

print(f"Total submissions to make: {len(data)}")

# === Setup ChromeDriver ===
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(), options=options)

# === Google Form URL ===
form_url = "https://forms.gle/m9dSToWpj5RGf4Kb8"

# === Fill Google Form ===
def fill_google_form(row_data):
    driver.get(form_url)
    time.sleep(3)

    questions = driver.find_elements(By.CSS_SELECTOR, "div[role='listitem']")

    for q in questions:
        try:
            question_text = q.text.split("\n")[0].strip()
            best_match, score = process.extractOne(question_text, headers)

            if score >= 70:
                answer = row_data[best_match]

                try:
                    input_box = q.find_element(By.CSS_SELECTOR, 'input[type="text"]')
                    input_box.send_keys(answer)
                except:
                    try:
                        text_area = q.find_element(By.CSS_SELECTOR, 'textarea')
                        text_area.send_keys(answer)
                    except:
                        print(f"❌ Could not fill: {question_text}")
            else:
                print(f"⚠️ No match for form question: {question_text}")
        except Exception as e:
            print(f"⚠️ Error matching question: {e}")

    try:
        submit_btn = driver.find_element(By.XPATH, "//span[contains(text(),'Submit')]/ancestor::div[@role='button']")
        submit_btn.click()
        print("✅ Form submitted.")
        time.sleep(2)
    except Exception as e:
        print("❌ Failed to submit form:", e)

# === Submit each row ===
for index, row in enumerate(data):
    print(f"\n🚀 Submitting row {index + 1}")
    fill_google_form(row)
    if index < len(data) - 1:
        print("⏳ Waiting 79 seconds before next submission...")
        time.sleep(79)

driver.quit()
